from odoo import models, fields, api
from odoo.exceptions import UserError
import base64
import io


class RepairMachineReportWizard(models.TransientModel):
    _name = 'repair.machine.report.wizard'
    _description = 'Wizard — Relatórios de OS'

    # ── Filtros de estado ─────────────────────────────────────────────

    state_ready      = fields.Boolean(string='Pronto',          default=True)
    state_progress   = fields.Boolean(string='Em Andamento',    default=True)
    state_paused     = fields.Boolean(string='Pausado',         default=False)
    state_pending_cq = fields.Boolean(string='Aguardando CQ',   default=False)
    state_done       = fields.Boolean(string='Concluído',       default=False)
    state_cancel     = fields.Boolean(string='Cancelado',       default=False)

    # ── Filtros opcionais — Programação ───────────────────────────────

    machine_id = fields.Many2one(
        comodel_name='repair.machine',
        string='Máquina',
        help='Deixe vazio para todas as máquinas.',
    )
    os_id = fields.Many2one(
        comodel_name='repair.order',
        string='Nº OS',
        help='Opcional. Filtra processos de uma OS específica.',
        domain=[('os_state', 'not in', ['cancel'])],
    )

    # ── Filtro de período ─────────────────────────────────────────────

    use_date_filter = fields.Boolean(string='Filtrar por período?', default=False)
    date_from = fields.Date(string='De',  default=fields.Date.today)
    date_to   = fields.Date(string='Até', default=fields.Date.today)

    # ── Filtro exclusivo — Impressão de OS Completa ───────────────────

    os_id_print = fields.Many2one(
        comodel_name='repair.order',
        string='Nº OS',
        help='Selecione a OS para imprimir completa.',
        domain=[('os_state', 'not in', ['cancel'])],
    )

    # ── Validação ─────────────────────────────────────────────────────

    @api.constrains('date_from', 'date_to')
    def _check_dates(self):
        for rec in self:
            if rec.use_date_filter and rec.date_from and rec.date_to:
                if rec.date_from > rec.date_to:
                    raise UserError('A Data Inicial não pode ser maior que a Data Final.')

    # ── Ações ─────────────────────────────────────────────────────────

    def action_print_report(self):
        self.ensure_one()
        states = self._get_selected_states()
        if not states:
            raise UserError('Selecione pelo menos um Estado para o relatório.')
        return self.env.ref(
            'cylinder_repair_os.action_report_machine_schedule'
        ).report_action(self)

    def action_print_os(self):
        self.ensure_one()
        if not self.os_id_print:
            raise UserError('Selecione um Nº OS para imprimir.')
        return self.env.ref(
            'cylinder_repair_os.action_report_os'
        ).report_action(self.os_id_print)

    def action_export_xls(self):
        """Exporta processos filtrados em XLSX — tabela simples."""
        self.ensure_one()
        processes = self._get_processes()
        if not processes:
            raise UserError('Nenhum processo encontrado com os filtros selecionados.')

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise UserError('Biblioteca openpyxl não disponível no servidor.')

        # Prefetch para evitar N+1
        processes.mapped('repair_id.partner_id')
        processes.mapped('repair_id.cylinder_id')
        processes.mapped('machine_id')
        processes.mapped('operator_id')
        processes.mapped('component_type_id')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Programação'

        state_labels = {
            'ready': 'Pronto', 'progress': 'Em Andamento',
            'paused': 'Pausado', 'pending_cq': 'Aguardando CQ',
            'done': 'Concluído', 'cancel': 'Cancelado',
        }

        headers = [
            'OS', 'Cliente', 'Componente', 'Operação', 'Máquina', 'Operador',
            'Data Prog.', 'Dt. Início', 'Dt. Conclusão',
            'Prev. (min)', 'Tempo Real (min)', 'Situação',
            'Pausas', 'Lead Time (min)', 'Eficiência (%)', 'Desvio',
        ]

        header_fill = PatternFill('solid', fgColor='4F46E5')
        header_font = Font(bold=True, color='FFFFFF')

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row, p in enumerate(processes, 2):
            ws.cell(row=row, column=1,  value=p.repair_id.os_number or '')
            ws.cell(row=row, column=2,  value=p.repair_id.partner_id.name if p.repair_id.partner_id else '')
            ws.cell(row=row, column=3,  value=p.component_type_id.name or '')
            ws.cell(row=row, column=4,  value=p.name or '')
            ws.cell(row=row, column=5,  value=p.machine_id.name if p.machine_id else '')
            ws.cell(row=row, column=6,  value=p.operator_id.name if p.operator_id else '')
            ws.cell(row=row, column=7,  value=str(p.date_planned) if p.date_planned else '')
            ws.cell(row=row, column=8,  value=str(p.date_start_orig)[:16] if p.date_start_orig else '')
            ws.cell(row=row, column=9,  value=str(p.date_finished)[:16] if p.date_finished else '')
            ws.cell(row=row, column=10, value=round(p.duration_planned or 0, 1))
            ws.cell(row=row, column=11, value=round(p.duration_acc or 0, 1))
            ws.cell(row=row, column=12, value=state_labels.get(p.state, p.state))
            ws.cell(row=row, column=13, value=p.pause_count or 0)
            ws.cell(row=row, column=14, value=round(p.lead_time_minutes or 0, 1))
            ws.cell(row=row, column=15, value=round(p.efficiency_pct or 0, 1))
            ws.cell(row=row, column=16, value='Sim' if p.has_deviation else 'Não')

        col_widths = [12, 20, 18, 25, 18, 18, 12, 16, 16, 12, 16, 16, 8, 14, 14, 8]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        attachment = self.env['ir.attachment'].create({
            'name': 'programacao_os.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(buf.read()).decode(),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%d?download=true' % attachment.id,
            'target': 'new',
        }

    # ── Helpers ───────────────────────────────────────────────────────

    def _get_selected_states(self):
        states = []
        if self.state_ready:      states.append('ready')
        if self.state_progress:   states.append('progress')
        if self.state_paused:     states.append('paused')
        if self.state_pending_cq: states.append('pending_cq')
        if self.state_done:       states.append('done')
        if self.state_cancel:     states.append('cancel')
        return states

    def _get_processes(self):
        self.ensure_one()
        states = self._get_selected_states()
        domain = [('state', 'in', states)] if states else [('id', '=', False)]

        if self.machine_id:
            domain.append(('machine_id', '=', self.machine_id.id))
        if self.os_id:
            domain.append(('repair_id', '=', self.os_id.id))

        if self.use_date_filter and self.date_from and self.date_to:
            domain += [
                '|',
                '&',
                ('date_planned', '>=', str(self.date_from)),
                ('date_planned', '<=', str(self.date_to)),
                '&',
                ('date_start_orig', '>=', '%s 00:00:00' % self.date_from),
                ('date_start_orig', '<=', '%s 23:59:59' % self.date_to),
            ]

        return self.env['repair.os.process'].search(
            domain,
            order='repair_id, component_type_id, sequence',
        )
